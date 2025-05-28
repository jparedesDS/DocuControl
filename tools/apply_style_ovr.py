import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Función para aplicar estilos y guardar el Excel
def aplicar_estilos_y_guardar_excel(df, filename):
    wb = Workbook()
    ws = wb.active

    # Estilos
    cell_filling_blue_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
    cell_filling = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
    medium_dashed = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    font_white = Font(color='FFFFFF', bold=True)
    font_black = Font(color='000000')
    font_black_bold = Font(color='000000', bold=True)

    estado_colors = {
        'Aprobado': PatternFill(start_color="00D25F", end_color="00D25F", fill_type="solid"),
        'Rechazado': PatternFill(start_color="FFA19A", end_color="FFA19A", fill_type="solid"),
        'Comentado': PatternFill(start_color="F79646", end_color="F79646", fill_type="solid"),
        'Com. Menores': PatternFill(start_color="FFE5AD", end_color="FFE5AD", fill_type="solid"),
        'Com. Mayores': PatternFill(start_color="DBB054", end_color="DBB054", fill_type="solid"),
        'Enviado': PatternFill(start_color="B1E1B9", end_color="B1E1B9", fill_type="solid"),
        'Sin Enviar': PatternFill(start_color="FFFFAB", end_color="FFFFAB", fill_type="solid"),
    }

    headers = list(df.columns)
    titulo_idx = headers.index('Título') + 1 if 'Título' in headers else None
    tag_original_idx = headers.index('TAG_ORIGINAL') + 1 if 'TAG_ORIGINAL' in headers else None

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)

        for c_idx, val in enumerate(row, start=1):
            cell_obj = ws.cell(row=r_idx, column=c_idx)
            cell_obj.border = medium_dashed

            if r_idx == 1:  # Encabezado
                cell_obj.fill = cell_filling
                cell_obj.font = font_white
            else:
                cell_obj.fill = cell_filling_blue_light
                cell_obj.font = font_black

                if isinstance(val, pd.Timestamp):
                    cell_obj.number_format = 'DD/MM/YYYY'

                # Colorear según estado
                if headers[c_idx - 1] == 'Estado':
                    estado = str(val).strip()
                    fill = estado_colors.get(estado, cell_filling_blue_light)
                    cell_obj.fill = fill
                    cell_obj.font = font_black_bold

                # Título en verde o gris si no tiene tag
                if titulo_idx and tag_original_idx and c_idx == titulo_idx and r_idx > 1:
                    tag_value = df.iloc[r_idx - 2, tag_original_idx - 1]
                    color = "00A249" if not pd.isna(tag_value) else "808080"
                    cell_obj.font = Font(color=color, bold=True)

    # Ajustar ancho
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "B2" # Fijar primera fila y primera columna
    wb.save(filename)