import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def aplicar_estilos_y_guardar_excel(df, filename):
    # Crear un nuevo libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active

    # Definir los estilos
    cell_filling_blue_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
    cell_filling = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
    medium_dashed = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    font_white = Font(color='FFFFFF', bold=True)
    font_black = Font(color='000000')

    # Aplicar borde a todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = medium_dashed

    # Convertir el DataFrame a filas de la hoja de Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        for c_idx, cell in enumerate(row, 1):
            if r_idx == 0:
                # Aplicar estilo a la cabecera
                cell_obj = ws.cell(row=r_idx+1, column=c_idx)
                cell_obj.fill = cell_filling
                cell_obj.font = font_white
                cell_obj.border = medium_dashed
            else:
                # Aplicar estilo a las celdas de datos
                cell_obj = ws.cell(row=r_idx+1, column=c_idx)
                cell_obj.border = medium_dashed
                cell_obj.font = font_black
                cell_obj.fill = cell_filling_blue_light
                if isinstance(cell, pd.Timestamp):
                    cell_obj.fill = cell_filling_blue_light

        # Ajustar ancho de columna al contenido
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Aplicar autofiltro para la primera fila
    ws.auto_filter.ref = ws.dimensions

    # Guardar el archivo Excel
    wb.save(filename)


def aplicar_estilos_html(df):
    styles = {
        'fecha': 'background-color: #D4DCF4; text-align: left; font-size: 14px;',
        'header': 'background-color: #6678AF; color: #FFFFFF; text-align: left; font-size: 14px;',
        'cell_even': 'background-color: #D4DCF4; text-align: left; font-size: 14px;',
        'cell_default': 'background-color: #D4DCF4; text-align: left; font-size: 14px;'
    }

    def style_specific_cell(val):
        if isinstance(val, pd.Timestamp):
            return styles['fecha']
        return styles['cell_even']

    def apply_conditional_styles(val):
        if val == 'Rechazado':
            return 'color: #000000; font-weight: bold; background-color: #FFA19A; font-size: 14px;'
        elif val == 'Com. Menores':
            return 'color: #000000; font-weight: bold; background-color: #FFE5AD; font-size: 14px;'
        elif val == 'Com. Mayores':
            return 'color: #000000; font-weight: bold; background-color: #DBB054; font-size: 14px;'
        elif val == 'Comentado':
            return 'color: #000000; font-weight: bold; background-color: #F79646; font-size: 14px;'
        elif val == 'Aprobado':
            return 'color: #000000; font-weight: bold; background-color: #00D25F; font-size: 14px;'
        elif val == 'Eliminado':
            return 'color: #000000; font-weight: bold; background-color: #FF0000; font-size: 14px;'
        else:
            return 'text-align: left; font-size: 14px;'

    header_style = [{'selector': 'th', 'props': [('background-color', '#6678AF'),
                                                 ('color', '#FFFFFF'),
                                                 ('text-align', 'center'),
                                                 ('font-size', '14px'),
                                                 ('font-weight', 'bold')]}]

    # Aplicar los estilos con .map en lugar de .applymap
    styled = df.style \
        .map(style_specific_cell) \
        .map(apply_conditional_styles) \
        .set_table_styles(header_style)

    # Convertir a HTML sin índice
    return styled.to_html(index=False)


def aplicar_estilo_info(df):
    # Aplicar estilos básicos a todas las celdas
    estilo_celdas = 'background-color: #D4DCF4; text-align: left; font-size: 14px;'
    estilo_header = [{'selector': 'th', 'props': [('background-color', '#6678AF'),
                                                  ('color', '#FFFFFF'),
                                                  ('text-align', 'center'),
                                                  ('font-size', '15px'),
                                                  ('font-weight', 'bold')]}]

    styled = df.style \
        .map(lambda _: estilo_celdas) \
        .set_table_styles(estilo_header)

    return styled.to_html(index=False)