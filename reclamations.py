import pandas as pd
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from copy import copy
import sys
import pandas as pd

today_date = pd.to_datetime('today', format='%d-%m-%Y', dayfirst=True)  # Capturamos la fecha actual del día
today_date_str = today_date.strftime('%d-%m-%Y') # Formateamos la fecha_actual a strf para la lectura y guardado de archivos

# Ruta del archivo original
ruta_excel = "C:\\Users\\alejandro.berzal\\Desktop\\DATA SCIENCE\\DocuControl\\Monitoring_Report_" + str(today_date_str) + ".xlsx"

# Cargar DataFrame de 'ENVIADOS'
df = pd.read_excel(ruta_excel, sheet_name='ENVIADOS')

# Filtrar filas con 15 o más días
df_filtrado = df[df['Días Devolución'] >= 0]

# Limpiar nombre de hoja
def limpiar_nombre(nombre):
    nombre = re.sub(r'[\\/*?:[\]]', '-', str(nombre))
    return nombre[:31]

# Cargar el libro original para extraer estilos
wb_original = load_workbook(ruta_excel)
ws_original = wb_original['ENVIADOS']

# Crear nuevo libro
wb_nuevo = Workbook()
wb_nuevo.remove(wb_nuevo.active)

# Definir color rojo y negrita para el texto
rojo_negrita_font = Font(color='FF5B5B', bold=True)

# Identificar columna 'Días Devolución'
columna_dias = df.columns.get_loc('Días Devolución') + 1  # índice Excel (1-based)

# Crear hoja por pedido
for pedido, grupo in df_filtrado.groupby('Nº Pedido'):
    nombre_hoja = limpiar_nombre(str(pedido))
    ws_nueva = wb_nuevo.create_sheet(title=nombre_hoja)

    # Insertar encabezado
    encabezado = list(df.columns)
    for col_idx, valor in enumerate(encabezado, start=1):
        celda_origen = ws_original.cell(row=1, column=col_idx)
        celda_nueva = ws_nueva.cell(row=1, column=col_idx, value=valor)

        celda_nueva.font = copy(celda_origen.font)
        celda_nueva.fill = copy(celda_origen.fill)
        celda_nueva.border = copy(celda_origen.border)

    # Diccionario para almacenar el ancho máximo de cada columna
    anchos_columnas = {col: len(str(valor)) for col, valor in enumerate(encabezado, start=1)}

    # Insertar filas con estilos
    for fila_nueva_idx, fila_df_idx in enumerate(grupo.index, start=2):
        fila_excel_idx = fila_df_idx + 2  # sumar encabezado

        for col_idx, valor in enumerate(grupo.loc[fila_df_idx], start=1):
            celda_origen = ws_original.cell(row=fila_excel_idx, column=col_idx)
            celda_nueva = ws_nueva.cell(row=fila_nueva_idx, column=col_idx, value=valor)

            # Copiar estilo
            celda_nueva.font = copy(celda_origen.font)
            celda_nueva.fill = copy(celda_origen.fill)
            celda_nueva.border = copy(celda_origen.border)

            # Colorear el texto de 'Días Devolución' con rojo y negrita
            if col_idx == columna_dias:
                try:
                    if float(valor) >= 15:
                        celda_nueva.font = rojo_negrita_font
                except ValueError:
                    pass

            # Ajustar el ancho máximo de la columna
            anchos_columnas[col_idx] = max(anchos_columnas[col_idx], len(str(valor)))

    # Añadir autofiltro a la primera fila
    ws_nueva.auto_filter.ref = ws_nueva.dimensions

    # Ajustar el ancho de las columnas
    for col_idx, ancho in anchos_columnas.items():
        ws_nueva.column_dimensions[chr(64 + col_idx)].width = ancho + 3

# Guardar nuevo archivo
wb_nuevo.save("Reclamaciones_" + str(today_date_str) + ".xlsx")
