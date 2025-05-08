# main.py

# Monitoring report
import os
import time
import pandas as pd
import xlsxwriter
import jpype
import asposecells
jpype.startJVM()
from asposecells.api import Workbook
from openpyxl.reader.excel import load_workbook
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Font
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from sqlalchemy import create_engine

# Email-mapi-automation
import shutil
import win32com.client
from bs4 import BeautifulSoup
from tools import *
from io import StringIO
import re
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows


# Importar módulos desde la carpeta my_modules
from tools import *
